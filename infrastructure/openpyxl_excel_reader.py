from openpyxl import load_workbook
from typing import Generator, Tuple, List
from interfaces.excel_reader import IExcelReader

class OpenPyxlExcelReader(IExcelReader):
    def read_rows(self, filepath: str) -> Generator[Tuple[str, List], None, None]:
        """
        openpyxl を使用して Excel ファイルを読み取り、シート名と行データをジェネレーターで返す。
        """
        try:
            workbook = load_workbook(filepath, read_only=True)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows(values_only=True):
                    yield (sheet, list(row))
        except Exception as e:
            raise RuntimeError(f"Failed to read {filepath}: {e}")