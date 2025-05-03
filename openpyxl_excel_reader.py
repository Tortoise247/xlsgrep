# openpyxl_excel_reader.py
import os
from typing import List, Tuple
from openpyxl import load_workbook
from excel_reader import IExcelReader

class OpenpyxlExcelReader(IExcelReader):
    def read_rows(self, filepath: str) -> List[Tuple]:
        rows = []
        wb = load_workbook(filepath, data_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                rows.append((sheet.title, row))
        wb.close()
        return rows
