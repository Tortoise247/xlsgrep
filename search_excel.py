import os
from openpyxl import load_workbook

def search_excel(directory, keyword):
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx'):
                filepath = os.path.join(root, file)
                try:
                    wb = load_workbook(filepath, data_only=True)
                    for sheet in wb.worksheets:
                        for row in sheet.iter_rows(values_only=True):
                            for cell in row:
                                if isinstance(cell, str) and keyword in cell:
                                    print(f"{filepath} | {sheet.title}: {cell}")
                    wb.close()
                except Exception as e:
                    print(f"Error opening {filepath}: {e}")

if __name__ == "__main__":
    dir_path = input("検索するフォルダ: ")
    keyword = input("検索するキーワード: ")
    search_excel(dir_path, keyword)
